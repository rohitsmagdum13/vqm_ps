"""Module: services/portal_intake/text_extractor.py

Text extraction for portal-uploaded attachments.

Strategy:
  - PDFs / images: try AWS Textract first (sync detect_document_text
    against the file already uploaded to S3). On Textract failure
    (access denied, multi-page, unsupported, throttling, etc.) fall
    back to pdfplumber for PDFs; images record extraction_status=failed.
  - DOCX: python-docx. Textract does not accept Word documents.
  - XLSX / XLS: openpyxl. Textract does not accept spreadsheets.
  - CSV / TXT: UTF-8 decode.
  - Anything else: not extracted.

The caller (AttachmentProcessor) supplies both the bytes (for fallback
parsing) and the S3 location (for Textract). Returning the method used
lets downstream consumers know whether OCR was involved.
"""

from __future__ import annotations

import asyncio
from pathlib import PurePosixPath
from typing import Literal

import structlog

from adapters.textract import TextractConnector, TextractError

logger = structlog.get_logger(__name__)

# Truncation limit shared with the email path — keeps token cost predictable
# downstream when the extracted text is fed into the entity-extraction prompt.
MAX_EXTRACTED_TEXT_LENGTH = 5000

ExtractionMethod = Literal[
    "textract", "pdfplumber", "openpyxl", "python_docx", "decode", "none"
]

# Content types / extensions that AWS Textract sync accepts.
_TEXTRACT_PDF_EXTS: frozenset[str] = frozenset({".pdf"})
_TEXTRACT_IMAGE_EXTS: frozenset[str] = frozenset({".png", ".jpg", ".jpeg", ".tiff", ".tif"})


class TextExtractor:
    """Extracts text from portal attachments using Textract first, then library fallbacks."""

    def __init__(self, textract: TextractConnector | None) -> None:
        """Initialize with an optional Textract connector.

        Textract is optional — when None (e.g. office IAM lacks the
        permission), PDFs go straight to pdfplumber and image files
        cannot be OCR'd.
        """
        self._textract = textract

    async def extract(
        self,
        *,
        content: bytes,
        content_type: str,
        filename: str,
        bucket: str,
        s3_key: str,
        correlation_id: str,
    ) -> tuple[str | None, ExtractionMethod]:
        """Return (text, method_used).

        text is None when extraction failed or was skipped. method_used
        is one of textract / pdfplumber / openpyxl / python_docx / decode
        / none and is recorded on the QueryAttachment row.
        """
        ext = PurePosixPath(filename).suffix.lower()

        # Textract-eligible files (PDF + images) — try OCR first.
        if ext in _TEXTRACT_PDF_EXTS or ext in _TEXTRACT_IMAGE_EXTS:
            if self._textract is not None:
                try:
                    text = await self._textract.detect_text_from_s3(
                        bucket=bucket,
                        key=s3_key,
                        correlation_id=correlation_id,
                    )
                    if text:
                        return text[:MAX_EXTRACTED_TEXT_LENGTH], "textract"
                except TextractError:
                    # Logged inside the adapter — fall through to fallback.
                    pass

            # Fallback for PDFs only — images have no library fallback.
            if ext in _TEXTRACT_PDF_EXTS:
                text = await asyncio.to_thread(self._extract_pdf_text, content, filename)
                if text:
                    return text[:MAX_EXTRACTED_TEXT_LENGTH], "pdfplumber"
            return None, "none"

        if ext in (".xlsx", ".xls") or "spreadsheet" in content_type:
            text = await asyncio.to_thread(self._extract_excel_text, content, filename)
            if text:
                return text[:MAX_EXTRACTED_TEXT_LENGTH], "openpyxl"
            return None, "none"

        if ext == ".docx" or "wordprocessingml" in content_type:
            text = await asyncio.to_thread(self._extract_docx_text, content, filename)
            if text:
                return text[:MAX_EXTRACTED_TEXT_LENGTH], "python_docx"
            return None, "none"

        if ext in (".csv", ".txt") or content_type.startswith("text/"):
            try:
                decoded = content.decode("utf-8", errors="replace")
            except Exception:
                logger.warning("Decode failed", file_name=filename)
                return None, "none"
            return decoded[:MAX_EXTRACTED_TEXT_LENGTH], "decode"

        return None, "none"

    @staticmethod
    def _extract_pdf_text(content: bytes, filename: str) -> str | None:
        """Extract text from PDF bytes using pdfplumber. Returns None on failure."""
        try:
            import io

            import pdfplumber

            text_parts: list[str] = []
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)
            return "\n".join(text_parts)
        except Exception:
            logger.warning("pdfplumber extraction failed", file_name=filename)
            return None

    @staticmethod
    def _extract_excel_text(content: bytes, filename: str) -> str | None:
        """Extract text from Excel bytes using openpyxl. Returns None on failure."""
        try:
            import io

            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            text_parts: list[str] = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        text_parts.append(row_text)
            wb.close()
            return "\n".join(text_parts)
        except Exception:
            logger.warning("openpyxl extraction failed", file_name=filename)
            return None

    @staticmethod
    def _extract_docx_text(content: bytes, filename: str) -> str | None:
        """Extract text from Word document bytes using python-docx. Returns None on failure."""
        try:
            import io

            from docx import Document

            doc = Document(io.BytesIO(content))
            text_parts = [para.text for para in doc.paragraphs if para.text.strip()]
            return "\n".join(text_parts)
        except Exception:
            logger.warning("python-docx extraction failed", file_name=filename)
            return None
