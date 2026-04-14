"""Module: services/email_intake.py

Email Ingestion Service for VQMS.

Handles the full 10-step email ingestion pipeline: fetch from
Graph API, idempotency check, parse fields, store raw email in S3,
process attachments, identify vendor via Salesforce, correlate
threads, generate IDs, write to database, publish events, and
enqueue to SQS for the AI pipeline.

Corresponds to Steps E1-E2.9 in the VQMS Solution Flow Document.

Usage:
    service = EmailIntakeService(graph_api, postgres, s3, sqs, eventbridge, salesforce, settings)
    result = await service.process_email("AAMkAGI2...", correlation_id="abc-123")
"""

from __future__ import annotations

import base64
import re
from pathlib import PurePosixPath

import structlog

from config.s3_paths import (
    FILENAME_RAW_EMAIL,
    S3_PREFIX_ATTACHMENTS,
    S3_PREFIX_INBOUND_EMAILS,
    build_s3_key,
)
from config.settings import Settings
from events.eventbridge import EventBridgeConnector
from adapters.graph_api import GraphAPIConnector
from storage.s3_client import S3Connector
from adapters.salesforce import SalesforceConnector
from queues.sqs import SQSConnector
from services.attachment_manifest import AttachmentManifestBuilder
from models.email import EmailAttachment, ParsedEmailPayload
from models.query import UnifiedQueryPayload
from utils.decorators import log_service_call
from utils.helpers import IdGenerator, TimeHelper

logger = structlog.get_logger(__name__)


class EmailIntakeError(Exception):
    """Raised when email ingestion fails at a critical step."""


class EmailIntakeService:
    """Handles the full email ingestion pipeline.

    Implements the 10-step email ingestion flow with critical
    vs non-critical step classification. Critical steps propagate
    errors (SQS retries the message). Non-critical steps log
    warnings and continue with safe defaults.
    """

    # Safety guardrails for attachments
    BLOCKED_EXTENSIONS: frozenset[str] = frozenset({".exe", ".bat", ".cmd", ".ps1", ".sh", ".js"})
    MAX_ATTACHMENT_SIZE: int = 10 * 1024 * 1024  # 10 MB per file
    MAX_TOTAL_SIZE: int = 50 * 1024 * 1024  # 50 MB per email
    MAX_ATTACHMENT_COUNT: int = 10
    MAX_EXTRACTED_TEXT_LENGTH: int = 5000

    def __init__(
        self,
        graph_api: GraphAPIConnector,
        postgres: object,  # PostgresConnector (typed as object to avoid circular import)
        s3: S3Connector,
        sqs: SQSConnector,
        eventbridge: EventBridgeConnector,
        salesforce: SalesforceConnector,
        settings: Settings,
    ) -> None:
        """Initialize with all required connectors.

        Args:
            graph_api: Microsoft Graph API connector for email fetch.
            postgres: PostgreSQL connector for metadata and idempotency.
            s3: S3 connector for raw email and attachment storage.
            sqs: SQS connector for enqueueing to the AI pipeline.
            eventbridge: EventBridge connector for event publishing.
            salesforce: Salesforce connector for vendor identification.
            settings: Application settings.
        """
        self._graph_api = graph_api
        self._postgres = postgres
        self._s3 = s3
        self._sqs = sqs
        self._eventbridge = eventbridge
        self._salesforce = salesforce
        self._settings = settings

    @log_service_call
    async def process_email(
        self,
        message_id: str,
        *,
        correlation_id: str | None = None,
    ) -> ParsedEmailPayload | None:
        """Process a single email through the full ingestion pipeline.

        This is the main entry point. It handles the 10-step flow:
        E2.1 Idempotency check, E1 Fetch email, E2.2 Parse fields,
        E2.7 Generate IDs, E2.3 Store raw in S3, E2.4 Process attachments,
        E2.5 Vendor identification, E2.6 Thread correlation,
        E2.8 Write to DB, E2.9a EventBridge, E2.9b SQS enqueue.

        Args:
            message_id: Exchange Online message ID to process.
            correlation_id: Tracing ID. Generated if not provided.

        Returns:
            ParsedEmailPayload with all extracted fields, or None
            if the email was a duplicate (already processed).

        Raises:
            EmailIntakeError: If a critical step fails.
        """
        correlation_id = correlation_id or IdGenerator.generate_correlation_id()

        # Bind correlation_id to structlog contextvars so all downstream
        # log calls (including in connectors) automatically include it
        structlog.contextvars.bind_contextvars(correlation_id=correlation_id)

        # E2.1 [CRITICAL] Idempotency check — prevent duplicate processing
        is_new = await self._postgres.check_idempotency(
            message_id, "email", correlation_id
        )
        if not is_new:
            logger.info(
                "Duplicate email skipped",
                message_id=message_id,
                correlation_id=correlation_id,
            )
            return None

        # E1 [CRITICAL] Fetch email from Graph API
        raw_email = await self._graph_api.fetch_email(
            message_id, correlation_id=correlation_id
        )

        # E2.2 [CRITICAL] Parse email fields
        parsed = self._parse_email_fields(raw_email)

        # E2.7 [CRITICAL] Generate IDs
        query_id = IdGenerator.generate_query_id()
        execution_id = IdGenerator.generate_execution_id()
        now = TimeHelper.ist_now()

        # E2.3 [NON-CRITICAL] Store raw email in S3
        s3_raw_key = await self._store_raw_email(raw_email, query_id, correlation_id)

        # E2.4 [NON-CRITICAL] Process attachments
        attachments = await self._process_attachments(raw_email, query_id, correlation_id)

        # E2.5 [NON-CRITICAL] Vendor identification via Salesforce
        vendor_id, vendor_match_method = await self._identify_vendor(parsed, correlation_id)

        # E2.6 [NON-CRITICAL] Thread correlation
        thread_status = await self._determine_thread_status(raw_email, correlation_id)

        # E2.8 [CRITICAL] Write metadata to database
        await self._store_email_metadata(
            message_id=message_id,
            query_id=query_id,
            correlation_id=correlation_id,
            parsed=parsed,
            s3_raw_key=s3_raw_key,
            vendor_id=vendor_id,
            vendor_match_method=vendor_match_method,
            thread_status=thread_status,
            now=now,
        )
        # Write attachment metadata to intake.email_attachments
        # Must happen AFTER email_messages INSERT because of the foreign key
        await self._store_attachment_metadata(
            message_id=message_id,
            query_id=query_id,
            attachments=attachments,
            correlation_id=correlation_id,
        )
        await self._create_case_execution(
            query_id=query_id,
            correlation_id=correlation_id,
            execution_id=execution_id,
            vendor_id=vendor_id,
            now=now,
        )

        # E2.9a [NON-CRITICAL] Publish EventBridge event
        try:
            await self._eventbridge.publish_event(
                "EmailParsed",
                {
                    "query_id": query_id,
                    "message_id": message_id,
                    "sender_email": parsed["sender_email"],
                    "vendor_id": vendor_id,
                },
                correlation_id=correlation_id,
            )
        except Exception:
            logger.warning(
                "EventBridge publish failed — continuing",
                query_id=query_id,
                correlation_id=correlation_id,
            )

        # E2.9b [CRITICAL] Enqueue to SQS for AI pipeline
        # Strip HTML body from email text for the pipeline payload
        body_text = parsed.get("body_text", "") or parsed.get("body_preview", "")
        payload = UnifiedQueryPayload(
            query_id=query_id,
            correlation_id=correlation_id,
            execution_id=execution_id,
            source="email",
            vendor_id=vendor_id,
            subject=parsed["subject"],
            body=body_text,
            priority="MEDIUM",  # Default; Query Analysis Agent sets final priority
            received_at=now,
            attachments=attachments,
            thread_status=thread_status,
            metadata={
                "message_id": message_id,
                "sender_email": parsed["sender_email"],
                "sender_name": parsed.get("sender_name"),
                "vendor_match_method": vendor_match_method,
                "conversation_id": parsed.get("conversation_id"),
            },
        )
        await self._sqs.send_message(
            self._settings.sqs_email_intake_queue_url,
            payload.model_dump(mode="json"),
            correlation_id=correlation_id,
        )

        # Build the full parsed payload for the return value
        result = ParsedEmailPayload(
            message_id=message_id,
            correlation_id=correlation_id,
            query_id=query_id,
            sender_email=parsed["sender_email"],
            sender_name=parsed.get("sender_name"),
            recipients=parsed.get("recipients", []),
            subject=parsed["subject"],
            body_text=body_text,
            body_html=parsed.get("body_html"),
            received_at=now,
            parsed_at=TimeHelper.ist_now(),
            in_reply_to=parsed.get("in_reply_to"),
            references=parsed.get("references", []),
            conversation_id=parsed.get("conversation_id"),
            thread_status=thread_status,
            vendor_id=vendor_id,
            vendor_match_method=vendor_match_method,
            attachments=attachments,
            s3_raw_email_key=s3_raw_key,
        )

        logger.info(
            "Email processed successfully",
            query_id=query_id,
            message_id=message_id,
            vendor_id=vendor_id,
            thread_status=thread_status,
            attachment_count=len(attachments),
            correlation_id=correlation_id,
        )
        return result

    # --- Private helper methods ---

    def _parse_email_fields(self, raw_email: dict) -> dict:
        """Extract structured fields from a Graph API message response.

        Pulls sender, recipients, subject, body, conversation ID,
        and reply-to headers from the Graph API response format.
        """
        from_field = raw_email.get("from", {}).get("emailAddress", {})
        sender_email = from_field.get("address", "unknown@unknown.com")
        sender_name = from_field.get("name")

        # Extract recipients
        recipients = []
        for r in raw_email.get("toRecipients", []):
            addr = r.get("emailAddress", {}).get("address", "")
            if addr:
                recipients.append(addr)

        # Extract body
        body_obj = raw_email.get("body", {})
        body_html = body_obj.get("content", "")
        body_text = self._html_to_text(body_html)
        body_preview = raw_email.get("bodyPreview", "")

        # Extract headers for thread correlation
        in_reply_to = ""
        references_list: list[str] = []
        for header in raw_email.get("internetMessageHeaders", []):
            name = header.get("name", "")
            value = header.get("value", "")
            if name == "In-Reply-To":
                in_reply_to = value
            elif name == "References":
                references_list = [ref.strip() for ref in value.split() if ref.strip()]

        return {
            "sender_email": sender_email,
            "sender_name": sender_name,
            "recipients": recipients,
            "subject": raw_email.get("subject", ""),
            "body_html": body_html,
            "body_text": body_text or body_preview,
            "body_preview": body_preview,
            "conversation_id": raw_email.get("conversationId"),
            "in_reply_to": in_reply_to or None,
            "references": references_list,
        }

    @staticmethod
    def _html_to_text(html: str) -> str:
        """Convert HTML to plain text by stripping tags.

        Simple regex-based approach for development. In production,
        consider using beautifulsoup4 for more robust parsing.
        """
        if not html:
            return ""
        # Remove HTML tags
        text = re.sub(r"<[^>]+>", " ", html)
        # Collapse multiple whitespace
        text = re.sub(r"\s+", " ", text).strip()
        return text

    async def _store_raw_email(
        self, raw_email: dict, query_id: str, correlation_id: str
    ) -> str | None:
        """Store the raw email JSON in S3. Non-critical — returns None on failure.

        Uses the single-bucket architecture: all files go to
        s3_bucket_data_store with prefix-based organization.
        Key pattern: inbound-emails/VQ-YYYY-NNNN/raw_email.json
        """
        try:
            import orjson

            raw_bytes = orjson.dumps(raw_email)
            key = build_s3_key(S3_PREFIX_INBOUND_EMAILS, query_id, FILENAME_RAW_EMAIL)
            await self._s3.upload_file(
                bucket=self._settings.s3_bucket_data_store,
                key=key,
                body=raw_bytes,
                content_type="application/json",
                correlation_id=correlation_id,
            )
            return key
        except Exception:
            logger.warning(
                "Failed to store raw email in S3 — continuing",
                query_id=query_id,
                correlation_id=correlation_id,
            )
            return None

    async def _process_attachments(
        self, raw_email: dict, query_id: str, correlation_id: str
    ) -> list[EmailAttachment]:
        """Process all attachments: validate, store to S3, extract text.

        Uses the single-bucket architecture: all attachments go to
        s3_bucket_data_store with prefix-based organization.
        Key pattern: attachments/VQ-YYYY-NNNN/{att_id}_{filename}

        After processing, stores a _manifest.json summarizing all
        attachments for this query.

        Non-critical — returns empty list on total failure, or
        partial results if some attachments fail.
        """
        raw_attachments = raw_email.get("attachments", [])
        if not raw_attachments:
            return []

        results: list[EmailAttachment] = []
        total_size = 0

        for i, att in enumerate(raw_attachments[: self.MAX_ATTACHMENT_COUNT]):
            filename = att.get("name", f"attachment_{i}")
            size_bytes = att.get("size", 0)
            content_type = att.get("contentType", "application/octet-stream")
            att_id = att.get("id", f"ATT-{i:03d}")

            # Check blocked extensions
            ext = PurePosixPath(filename).suffix.lower()
            if ext in self.BLOCKED_EXTENSIONS:
                logger.warning(
                    "Blocked attachment extension",
                    file_name=filename,
                    ext=ext,
                    correlation_id=correlation_id,
                )
                results.append(
                    EmailAttachment(
                        attachment_id=att_id,
                        filename=filename,
                        content_type=content_type,
                        size_bytes=size_bytes,
                        extraction_status="skipped",
                    )
                )
                continue

            # Check individual file size
            if size_bytes > self.MAX_ATTACHMENT_SIZE:
                logger.warning(
                    "Oversized attachment skipped",
                    file_name=filename,
                    size_bytes=size_bytes,
                    correlation_id=correlation_id,
                )
                results.append(
                    EmailAttachment(
                        attachment_id=att_id,
                        filename=filename,
                        content_type=content_type,
                        size_bytes=size_bytes,
                        extraction_status="skipped",
                    )
                )
                continue

            # Check total size
            total_size += size_bytes
            if total_size > self.MAX_TOTAL_SIZE:
                logger.warning(
                    "Total attachment size exceeded — skipping remaining",
                    total_size=total_size,
                    correlation_id=correlation_id,
                )
                break

            # Try to decode and store the attachment
            try:
                content_bytes_b64 = att.get("contentBytes", "")
                content_bytes = base64.b64decode(content_bytes_b64) if content_bytes_b64 else b""

                # Store to S3 using single-bucket architecture
                s3_key = build_s3_key(
                    S3_PREFIX_ATTACHMENTS, query_id, f"{att_id}_{filename}"
                )
                await self._s3.upload_file(
                    bucket=self._settings.s3_bucket_data_store,
                    key=s3_key,
                    body=content_bytes,
                    content_type=content_type,
                    correlation_id=correlation_id,
                )

                # Extract text based on content type
                extracted_text = await self._extract_text(content_bytes, content_type, filename)

                results.append(
                    EmailAttachment(
                        attachment_id=att_id,
                        filename=filename,
                        content_type=content_type,
                        size_bytes=size_bytes,
                        s3_key=s3_key,
                        extracted_text=extracted_text,
                        extraction_status="success" if extracted_text else "failed",
                    )
                )
            except Exception:
                logger.warning(
                    "Attachment processing failed",
                    file_name=filename,
                    correlation_id=correlation_id,
                )
                results.append(
                    EmailAttachment(
                        attachment_id=att_id,
                        filename=filename,
                        content_type=content_type,
                        size_bytes=size_bytes,
                        extraction_status="failed",
                    )
                )

        # Store attachment manifest (non-critical)
        await AttachmentManifestBuilder.store_manifest(
            s3=self._s3,
            settings=self._settings,
            query_id=query_id,
            attachments=results,
            correlation_id=correlation_id,
        )

        return results

    async def _extract_text(
        self, content: bytes, content_type: str, filename: str
    ) -> str | None:
        """Extract text from an attachment based on its type.

        Supports PDF (pdfplumber), Excel (openpyxl), Word (python-docx),
        and CSV/TXT (direct decode). All extraction runs in a thread
        to avoid blocking the event loop.

        Returns text truncated to MAX_EXTRACTED_TEXT_LENGTH chars, or None.
        """
        import asyncio

        try:
            ext = PurePosixPath(filename).suffix.lower()

            if ext == ".pdf" or content_type == "application/pdf":
                text = await asyncio.to_thread(self._extract_pdf_text, content)
            elif ext in (".xlsx", ".xls") or "spreadsheet" in content_type:
                text = await asyncio.to_thread(self._extract_excel_text, content)
            elif ext == ".docx" or "wordprocessingml" in content_type:
                text = await asyncio.to_thread(self._extract_docx_text, content)
            elif ext in (".csv", ".txt") or content_type.startswith("text/"):
                text = content.decode("utf-8", errors="replace")
            else:
                # Unsupported type — skip extraction
                return None

            if text:
                return text[: self.MAX_EXTRACTED_TEXT_LENGTH]
            return None
        except Exception:
            logger.warning("Text extraction failed", file_name=filename)
            return None

    @staticmethod
    def _extract_pdf_text(content: bytes) -> str:
        """Extract text from PDF bytes using pdfplumber."""
        import io

        import pdfplumber

        text_parts = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        return "\n".join(text_parts)

    @staticmethod
    def _extract_excel_text(content: bytes) -> str:
        """Extract text from Excel bytes using openpyxl."""
        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        text_parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_parts.append(row_text)
        wb.close()
        return "\n".join(text_parts)

    @staticmethod
    def _extract_docx_text(content: bytes) -> str:
        """Extract text from Word document bytes using python-docx."""
        import io

        from docx import Document

        doc = Document(io.BytesIO(content))
        text_parts = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n".join(text_parts)

    async def _identify_vendor(
        self, parsed: dict, correlation_id: str
    ) -> tuple[str | None, str | None]:
        """Identify the vendor via Salesforce 3-step fallback.

        Non-critical — returns (None, "unresolved") on failure.
        """
        try:
            match = await self._salesforce.identify_vendor(
                sender_email=parsed["sender_email"],
                sender_name=parsed.get("sender_name"),
                body_text=parsed.get("body_text"),
                correlation_id=correlation_id,
            )
            if match is not None:
                return match.vendor_id, match.match_method
            return None, "unresolved"
        except Exception:
            logger.warning(
                "Vendor identification failed — continuing without vendor",
                sender=parsed["sender_email"],
                correlation_id=correlation_id,
            )
            return None, "unresolved"

    async def _determine_thread_status(
        self, raw_email: dict, correlation_id: str
    ) -> str:
        """Check if this email is part of an existing thread.

        Looks up the conversationId in workflow.case_execution
        to determine if this is a new query, a reply to an open
        case, or a reply to a closed case.

        Non-critical — returns "NEW" on failure.
        """
        conversation_id = raw_email.get("conversationId")
        if not conversation_id:
            return "NEW"

        try:
            row = await self._postgres.fetchrow(
                "SELECT query_id, status FROM workflow.case_execution "
                "WHERE conversation_id = $1 ORDER BY created_at DESC LIMIT 1",
                conversation_id,
            )
            if row is None:
                return "NEW"
            status = row.get("status", "")
            if status in ("CLOSED", "RESOLVED"):
                return "REPLY_TO_CLOSED"
            return "EXISTING_OPEN"
        except Exception:
            logger.warning(
                "Thread correlation failed — defaulting to NEW",
                conversation_id=conversation_id,
                correlation_id=correlation_id,
            )
            return "NEW"

    async def _store_email_metadata(
        self,
        *,
        message_id: str,
        query_id: str,
        correlation_id: str,
        parsed: dict,
        s3_raw_key: str | None,
        vendor_id: str | None,
        vendor_match_method: str | None,
        thread_status: str,
        now: object,
    ) -> None:
        """Write email metadata to intake.email_messages table."""
        await self._postgres.execute(
            """
            INSERT INTO intake.email_messages
            (message_id, query_id, correlation_id, sender_email, sender_name,
             subject, body_text, body_html, received_at, parsed_at,
             in_reply_to, conversation_id, thread_status, vendor_id,
             vendor_match_method, s3_raw_email_key, source, created_at)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10,
                    $11, $12, $13, $14, $15, $16, $17, $18)
            """,
            message_id,
            query_id,
            correlation_id,
            parsed["sender_email"],
            parsed.get("sender_name"),
            parsed["subject"],
            parsed.get("body_text", ""),
            parsed.get("body_html"),
            now,
            now,
            parsed.get("in_reply_to"),
            parsed.get("conversation_id"),
            thread_status,
            vendor_id,
            vendor_match_method,
            s3_raw_key,
            "email",
            now,
        )

    async def _store_attachment_metadata(
        self,
        *,
        message_id: str,
        query_id: str,
        attachments: list[EmailAttachment],
        correlation_id: str,
    ) -> None:
        """Write each attachment's metadata to intake.email_attachments.

        Non-critical — if a single attachment INSERT fails, log and
        continue with the rest. The pipeline can still proceed using
        the attachment data from the SQS payload.
        """
        if not attachments:
            return

        for att in attachments:
            try:
                await self._postgres.execute(
                    """
                    INSERT INTO intake.email_attachments
                    (message_id, query_id, attachment_id, filename, content_type,
                     size_bytes, s3_key, extracted_text, extraction_status)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                    ON CONFLICT (attachment_id) DO NOTHING
                    """,
                    message_id,
                    query_id,
                    att.attachment_id,
                    att.filename,
                    att.content_type,
                    att.size_bytes,
                    att.s3_key,
                    att.extracted_text[:5000] if att.extracted_text else None,
                    att.extraction_status,
                )
            except Exception:
                logger.warning(
                    "Failed to store attachment metadata — continuing",
                    attachment_id=att.attachment_id,
                    filename=att.filename,
                    query_id=query_id,
                    correlation_id=correlation_id,
                )

        logger.info(
            "Attachment metadata stored",
            query_id=query_id,
            attachment_count=len(attachments),
            correlation_id=correlation_id,
        )

    async def _create_case_execution(
        self,
        *,
        query_id: str,
        correlation_id: str,
        execution_id: str,
        vendor_id: str | None,
        now: object,
    ) -> None:
        """Create a new case execution record in workflow.case_execution."""
        await self._postgres.execute(
            """
            INSERT INTO workflow.case_execution
            (query_id, correlation_id, execution_id, vendor_id, source, status,
             created_at, updated_at)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
            """,
            query_id,
            correlation_id,
            execution_id,
            vendor_id,
            "email",
            "RECEIVED",
            now,
            now,
        )
