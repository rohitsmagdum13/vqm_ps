"""Module: services/email_intake/attachment_processor.py

Attachment validation, S3 storage, and text extraction.

Processes email attachments: validates against safety guardrails,
stores binary content to S3, extracts text from PDF/Excel/Word/CSV
files, and stores a manifest summarizing all attachments.
"""

from __future__ import annotations

import asyncio
import base64
from pathlib import PurePosixPath

import structlog

from adapters.graph_api import GraphAPIConnector
from config.s3_paths import S3_PREFIX_ATTACHMENTS, build_s3_key
from config.settings import Settings
from models.email import EmailAttachment
from services.attachment_manifest import AttachmentManifestBuilder
from storage.s3_client import S3Connector

logger = structlog.get_logger(__name__)

# Graph API inlines attachment bytes only up to ~3 MB. Anything larger
# must be fetched via a second API call. We use the same threshold Graph
# uses so our fallback kicks in exactly when their inline delivery drops
# the bytes.
_INLINE_BYTES_THRESHOLD: int = 3 * 1024 * 1024


class AttachmentProcessor:
    """Validates, stores, and extracts text from email attachments.

    Uses the single-bucket S3 architecture with prefix-based
    organization. Enforces safety guardrails (blocked extensions,
    size limits, count limits). Falls back to a per-attachment Graph
    API download for files whose bytes are too large for inline
    delivery.
    """

    # Safety guardrails for attachments
    BLOCKED_EXTENSIONS: frozenset[str] = frozenset({".exe", ".bat", ".cmd", ".ps1", ".sh", ".js"})
    MAX_ATTACHMENT_SIZE: int = 10 * 1024 * 1024  # 10 MB per file
    MAX_TOTAL_SIZE: int = 50 * 1024 * 1024  # 50 MB per email
    MAX_ATTACHMENT_COUNT: int = 10
    MAX_EXTRACTED_TEXT_LENGTH: int = 5000

    def __init__(
        self,
        s3: S3Connector,
        settings: Settings,
        graph_api: GraphAPIConnector | None = None,
    ) -> None:
        """Initialize with S3 connector, settings, and Graph API adapter.

        Args:
            s3: S3 connector for attachment storage.
            settings: Application settings (bucket name).
            graph_api: Optional Graph API connector used to download
                attachments that exceed the inline-bytes threshold. When
                not provided, oversized attachments are recorded with
                ``extraction_status='failed'`` — unit tests that don't
                exercise attachments can omit it.
        """
        self._s3 = s3
        self._settings = settings
        self._graph_api = graph_api

    async def process_attachments(
        self,
        raw_email: dict,
        query_id: str,
        correlation_id: str,
        *,
        message_id: str | None = None,
    ) -> list[EmailAttachment]:
        """Process all attachments: validate, store to S3, extract text.

        Key pattern: attachments/VQ-YYYY-NNNN/{att_id}_{filename}

        After processing, stores a _manifest.json summarizing all
        attachments for this query.

        Non-critical — returns empty list on total failure, or
        partial results if some attachments fail.
        """
        raw_attachments = raw_email.get("attachments", [])
        if not raw_attachments:
            return []

        results: list[EmailAttachment] = []
        total_size = 0

        for i, att in enumerate(raw_attachments[: self.MAX_ATTACHMENT_COUNT]):
            filename = att.get("name", f"attachment_{i}")
            size_bytes = att.get("size", 0)
            content_type = att.get("contentType", "application/octet-stream")
            att_id = att.get("id", f"ATT-{i:03d}")

            # Check blocked extensions
            ext = PurePosixPath(filename).suffix.lower()
            if ext in self.BLOCKED_EXTENSIONS:
                logger.warning(
                    "Blocked attachment extension",
                    file_name=filename,
                    ext=ext,
                    correlation_id=correlation_id,
                )
                results.append(
                    EmailAttachment(
                        attachment_id=att_id,
                        filename=filename,
                        content_type=content_type,
                        size_bytes=size_bytes,
                        extraction_status="skipped",
                    )
                )
                continue

            # Check individual file size
            if size_bytes > self.MAX_ATTACHMENT_SIZE:
                logger.warning(
                    "Oversized attachment skipped",
                    file_name=filename,
                    size_bytes=size_bytes,
                    correlation_id=correlation_id,
                )
                results.append(
                    EmailAttachment(
                        attachment_id=att_id,
                        filename=filename,
                        content_type=content_type,
                        size_bytes=size_bytes,
                        extraction_status="skipped",
                    )
                )
                continue

            # Check total size
            total_size += size_bytes
            if total_size > self.MAX_TOTAL_SIZE:
                logger.warning(
                    "Total attachment size exceeded — skipping remaining",
                    total_size=total_size,
                    correlation_id=correlation_id,
                )
                break

            # Try to decode and store the attachment
            try:
                content_bytes = await self._resolve_attachment_bytes(
                    att=att,
                    message_id=message_id,
                    size_bytes=size_bytes,
                    filename=filename,
                    correlation_id=correlation_id,
                )
                if not content_bytes:
                    # No bytes could be resolved — record and move on.
                    results.append(
                        EmailAttachment(
                            attachment_id=att_id,
                            filename=filename,
                            content_type=content_type,
                            size_bytes=size_bytes,
                            extraction_status="failed",
                        )
                    )
                    continue

                # Store to S3 using single-bucket architecture
                s3_key = build_s3_key(
                    S3_PREFIX_ATTACHMENTS, query_id, f"{att_id}_{filename}"
                )
                await self._s3.upload_file(
                    bucket=self._settings.s3_bucket_data_store,
                    key=s3_key,
                    body=content_bytes,
                    content_type=content_type,
                    correlation_id=correlation_id,
                )

                # Extract text based on content type
                extracted_text = await self._extract_text(content_bytes, content_type, filename)

                results.append(
                    EmailAttachment(
                        attachment_id=att_id,
                        filename=filename,
                        content_type=content_type,
                        size_bytes=size_bytes,
                        s3_key=s3_key,
                        extracted_text=extracted_text,
                        extraction_status="success" if extracted_text else "failed",
                    )
                )
            except Exception:
                logger.warning(
                    "Attachment processing failed",
                    file_name=filename,
                    correlation_id=correlation_id,
                )
                results.append(
                    EmailAttachment(
                        attachment_id=att_id,
                        filename=filename,
                        content_type=content_type,
                        size_bytes=size_bytes,
                        extraction_status="failed",
                    )
                )

        # Store attachment manifest (non-critical)
        await AttachmentManifestBuilder.store_manifest(
            s3=self._s3,
            settings=self._settings,
            query_id=query_id,
            attachments=results,
            correlation_id=correlation_id,
        )

        return results

    async def _resolve_attachment_bytes(
        self,
        *,
        att: dict,
        message_id: str | None,
        size_bytes: int,
        filename: str,
        correlation_id: str,
    ) -> bytes:
        """Return the raw bytes of a single attachment.

        Graph API inlines ``contentBytes`` (Base64) only for files up to
        ~3 MB. Above that the field is missing and a separate GET call
        to ``/messages/{id}/attachments/{att_id}/$value`` is required.
        We call that fallback when inline bytes are absent and we have
        both a Graph adapter and the parent message_id.

        Returns empty bytes when no source is available — the caller
        then records the attachment with ``extraction_status='failed'``.
        """
        content_bytes_b64 = att.get("contentBytes", "")
        if content_bytes_b64:
            return base64.b64decode(content_bytes_b64)

        # Inline bytes missing — either > 3 MB or Graph chose not to
        # inline. Try the large-attachment endpoint if we can.
        att_id = att.get("id")
        if not att_id or not message_id or self._graph_api is None:
            logger.warning(
                "Attachment has no inline bytes and cannot be downloaded",
                file_name=filename,
                size_bytes=size_bytes,
                has_graph=self._graph_api is not None,
                has_message_id=message_id is not None,
                correlation_id=correlation_id,
            )
            return b""

        if size_bytes <= _INLINE_BYTES_THRESHOLD:
            # Unusual: small attachment with no contentBytes. Log once
            # and still attempt download — sometimes Graph omits bytes
            # for message drafts or resent items.
            logger.info(
                "Small attachment missing inline bytes — fetching via /$value",
                file_name=filename,
                size_bytes=size_bytes,
                correlation_id=correlation_id,
            )

        try:
            return await self._graph_api.download_large_attachment(
                message_id,
                att_id,
                correlation_id=correlation_id,
            )
        except Exception:
            logger.warning(
                "Large attachment download failed — recording as failed",
                file_name=filename,
                attachment_id=att_id,
                size_bytes=size_bytes,
                correlation_id=correlation_id,
            )
            return b""

    async def _extract_text(
        self, content: bytes, content_type: str, filename: str
    ) -> str | None:
        """Extract text from an attachment based on its type.

        Supports PDF (pdfplumber), Excel (openpyxl), Word (python-docx),
        and CSV/TXT (direct decode). Runs in a thread to avoid blocking.

        Returns text truncated to MAX_EXTRACTED_TEXT_LENGTH chars, or None.
        """
        try:
            ext = PurePosixPath(filename).suffix.lower()

            if ext == ".pdf" or content_type == "application/pdf":
                text = await asyncio.to_thread(self._extract_pdf_text, content)
            elif ext in (".xlsx", ".xls") or "spreadsheet" in content_type:
                text = await asyncio.to_thread(self._extract_excel_text, content)
            elif ext == ".docx" or "wordprocessingml" in content_type:
                text = await asyncio.to_thread(self._extract_docx_text, content)
            elif ext in (".csv", ".txt") or content_type.startswith("text/"):
                text = content.decode("utf-8", errors="replace")
            else:
                # Unsupported type — skip extraction
                return None

            if text:
                return text[: self.MAX_EXTRACTED_TEXT_LENGTH]
            return None
        except Exception:
            logger.warning("Text extraction failed", file_name=filename)
            return None

    @staticmethod
    def _extract_pdf_text(content: bytes) -> str:
        """Extract text from PDF bytes using pdfplumber."""
        import io

        import pdfplumber

        text_parts = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        return "\n".join(text_parts)

    @staticmethod
    def _extract_excel_text(content: bytes) -> str:
        """Extract text from Excel bytes using openpyxl."""
        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        text_parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_parts.append(row_text)
        wb.close()
        return "\n".join(text_parts)

    @staticmethod
    def _extract_docx_text(content: bytes) -> str:
        """Extract text from Word document bytes using python-docx."""
        import io

        from docx import Document

        doc = Document(io.BytesIO(content))
        text_parts = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n".join(text_parts)
